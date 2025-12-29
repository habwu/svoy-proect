import io
import requests
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponseRedirect, HttpResponseBadRequest, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.translation import gettext as _
from django.views import View
from django.views.generic import ListView
from openpyxl import Workbook
from io import BytesIO
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Frame, PageTemplate, FrameBreak
from petrovich.main import Petrovich
from petrovich.enums import Case, Gender
from django.template import Template, Context
from django.conf import settings
import re
from collections import defaultdict
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import SimpleDocTemplate, Paragraph, FrameBreak
from django.http import JsonResponse
from rest_framework.views import APIView

from classroom.models import *
from files.models import PDFTemplate, AgreementSettings
from main.models import *
from main.permissions import IsAdminUser
from register.models import RegisterAdmin, RegisterSend
from result.models import *
from users.models import User
from openpyxl.styles import PatternFill
import threading
from .utils import *
import pandas as pd
import logging


class ExcelClassroomAPIView(APIView):
    permission_classes = (IsAdminUser,)

    def get(self, request, Classroom_id):
        classroom = get_object_or_404(Classroom, id=Classroom_id, school=request.user.school)
        queryset = RegisterAdmin.objects.filter(
            child_admin__classrom__id=Classroom_id,
            is_deleted=False,
            school=request.user.school,
        )

        subject_order = [
            "Английский язык (4, 5-6, 7-8, 9-11)", "География (5, 6, 7, 8, 9, 10-11)", "Информатика (3, 4)",
            "Искусство (МХК) (5, 6, 7, 8, 9, 10, 11)", "История (5, 6, 7, 8, 9, 10-11)", "ИЗО (7, 8, 9)",
            "Литература (5, 6, 7, 8, 9, 10, 11)", "Музыка (5, 6, 7, 8)", "Немецкий язык (4, 5-6, 7-8, 9-11)",
            "Обществознание (5, 6, 7, 8, 9, 10, 11)", "ОБЗР (5, 6, 7, 8, 9, 10-11)", "Право (9, 10, 11)",
            "Психология (7-11)", "Русский язык (5, 6, 7, 8, 9, 10, 11)", "Труд (технология) (5-6, 7-8, 9, 10-11)",
            "Физика (5, 6)", "Физическая культура (5-6, 7-8, 9-11)", "Французский язык (4, 5-6, 7-8, 9-11)",
            "Экология (7, 8, 9, 10, 11)", "Экономика (7-9, 10-11)", "НШ: литературное чтение (4)",
            "НШ: окружающий мир (4)", "НШ: русский язык (4)"
        ]

        # получение всех предметов из БД
        all_subjects = Subject.objects.all().values_list('name', flat=True)
        all_subjects_lower = [subject.lower() for subject in all_subjects]
        remaining_subjects = [subject for subject in all_subjects_lower if subject not in subject_order]
        ordered_subjects = subject_order + [subject.title() for subject in remaining_subjects]

        # excel-файл
        wb = Workbook()
        ws = wb.active

        # заголовки
        headers = ["№", "Фамилия", "Имя", "Отчество", "Пол", "Дата рождения", "Класс"] + ordered_subjects + [
            "Кол-во заявлений"]
        ws.append(headers)

        # заполнение данных
        for idx, obj in enumerate(queryset, start=1):
            student = obj.child_admin
            row = [
                      idx,
                      student.last_name,
                      student.first_name,
                      student.surname,
                      student.gender,
                      student.birth_date,
                      f"{student.classroom.number}{student.classroom.letter}"
                  ] + [0] * len(ordered_subjects) + [1]
            ws.append(row)

            # генерация ответа
            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            filename = f'zayvki_{classroom.number}{classroom.letter}_class.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)

            return response


class ExcelAllAPIView(APIView):
    permission_classes = (IsAdminUser,)

    def get(self, request):
        queryset = RegisterAdmin.objects.filter(is_deleted=False, school=request.user.school)

        subject_order = [
            "Английский язык (4, 5-6, 7-8, 9-11)", "География (5, 6, 7, 8, 9, 10-11)", "Информатика (3, 4)",
            "Искусство (МХК) (5, 6, 7, 8, 9, 10, 11)", "История (5, 6, 7, 8, 9, 10-11)", "ИЗО (7, 8, 9)",
            "Литература (5, 6, 7, 8, 9, 10, 11)", "Музыка (5, 6, 7, 8)", "Немецкий язык (4, 5-6, 7-8, 9-11)",
            "Обществознание (5, 6, 7, 8, 9, 10, 11)", "ОБЗР (5, 6, 7, 8, 9, 10-11)", "Право (9, 10, 11)",
            "Психология (7-11)", "Русский язык (5, 6, 7, 8, 9, 10, 11)", "Труд (технология) (5-6, 7-8, 9, 10-11)",
            "Физика (5, 6)", "Физическая культура (5-6, 7-8, 9-11)", "Французский язык (4, 5-6, 7-8, 9-11)",
            "Экология (7, 8, 9, 10, 11)", "Экономика (7-9, 10-11)", "НШ: литературное чтение (4)",
            "НШ: окружающий мир (4)", "НШ: русский язык (4)"
        ]

        # Получение всех предметов из базы данных
        all_subjects = Subject.objects.all().values_list('name', flat=True)
        all_subjects_lower = [subject.lower() for subject in all_subjects]
        remaining_subjects = [subject for subject in all_subjects_lower if subject not in subject_order]
        ordered_subjects = subject_order + [subject.title() for subject in remaining_subjects]

        wb = Workbook()
        ws = wb.active

        headers = [
                      "№", "Фамилия", "Имя", "Отчество", "Пол", "Дата рождения", "Класс", "Гражданство", "ОВЗ",
                      "Наименование ОУ"
                  ] + ordered_subjects + ["Кол-во заявлений"]
        ws.append(headers)
        # Заполнение данными
        student_data = {}

        for obj in queryset:
            student = obj.child_admin
            if student.id not in student_data:
                student_data[student.id] = {
                    "last_name": student.last_name,
                    "first_name": student.first_name,
                    "surname": student.surname,
                    "gender": student.gender,
                    "birth_date": student.birth_date,
                    "class": f"{student.classroom.number}{student.classroom.letter}",
                    "subjects": {subject: 0 for subject in ordered_subjects}
                }

            subject_name = obj.olympiad_admin.subject.name.lower()
            if subject_name in ordered_subjects:
                student_data[student.id]["subjects"][subject_name] = 1

        for idx, (student_id, student_info) in enumerate(student_data.items(), start=1):
            row = [
                      idx,
                      student_info["last_name"],
                      student_info["first_name"],
                      student_info["surname"],
                      student_info["gender"],
                      student_info["birth_date"],
                      student_info["class"],
                      "РФ",  # Гражданство
                      "",  # ОВЗ
                      "МАОУ «МЛ № 1»"  # Наименование ОУ
                  ] + [student_info["subjects"][subject] for subject in ordered_subjects] + [
                      sum(student_info["subjects"].values())
                  ]
            ws.append(row)

        # Генерация ответа
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Baza_{request.user.school.name}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response



